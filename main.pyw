from gui import Ui_Form
from PyQt5 import QtCore, QtGui, QtWidgets, QtMultimedia
from openpyxl import load_workbook, Workbook
from datetime import timedelta
from plyer import notification
import re

style_sheet = """
    QTabBar::tab{
        background: rgb(43, 4, 66);
        color: black;
        border: 4px solid rgb(72, 5, 122);
        border-right-width: 0px;	
        border-radius: 10px 0;
        margin-top: 10px;
        margin-bottom: 10px;
        padding: 5px;
        padding-right: -8px
    }

    QTabBar::tab:selected  {
        background: rgb(72, 5, 122);
        color: gray;
        border-color: rgb(43, 4, 66);
    }

    QTabWidget>QWidget>QWidget{
        background: rgb(72, 5, 122);
        border: 1px solid white; border-left:0px
    }

    QCheckBox::indicator{
        width :30px;
        height :30px;
        border: 5px solid black;
        background: transparent
    }

    QCheckBox::indicator:unchecked:pressed{
        background-color : lightgreen;
    }

    QCheckBox::indicator:checked:pressed{
        background-color : #fd4e4e;
    }

    QCheckBox::indicator:checked{
        background-color : green;
    }

    
    QCheckBox::indicator:unchecked{
        background-color : red;
    }

    QRoundProgressBar {
        background-color: rgb(118, 52, 118)
        }

    QPushButton#start{
        background: transparent;
        border: 5px solid black;
        border-radius: 10px;
        color:white
    }

    QPushButton#start::hover{
        color: black
    }
    
    QComboBox#comboBox{
        background: rgb(61, 4, 94);
        border: 5px solid black;
        border-radius: 5px;
        font-size: 20px;
        }

    QComboBox#status {
        border: 1px solid black; font-size: 10px
    }
"""

class POMODORO(QtWidgets.QWidget):
    def __init__(self, MainWindow):
        super(POMODORO, self).__init__()
        self.MainWindow = MainWindow
        self.ui = Ui_Form()
        self.ui.setupUi(self.MainWindow)
        self.inintalizeUI()

    def inintalizeUI(self):
        self.update()
        self.ui.tableWidget.itemChanged.connect(self.edit)
        self.ui.pushButton.clicked.connect(self.counterOption)
        self.ui.task.setCurrentText('Select a task or, Enter new one.')
        self.ui.task.currentTextChanged.connect(self.taskOption)
        self.ui.close.clicked.connect(self.close)
        self.ui.minimize.clicked.connect(lambda: self.MainWindow.showMinimized())
        self.ui.status.clicked.connect(lambda: self.save('Done'))
        
        url = QtCore.QUrl.fromLocalFile('alert.wav')
        content = QtMultimedia.QMediaContent(url)
        self.alert = QtMultimedia.QMediaPlayer()
        self.alert.setMedia(content)
        self.item_text = None

        self.curent_time = 25 * 60
        self.ui.progressBar.setValue(0)
        self.ui.progressBar.setRange(0, 25 * 60)
        self.counter = 25 * 60
        self.passed_time = 25 * 60
        self.timer = QtCore.QTimer()
        self.loop = 1

        self.timer.timeout.connect(self.handleTimer)

    def close(self):
        message = QtWidgets.QMessageBox()
        answer = message.question(
            self, "Quit!", "Are your sure you want to exit the app?",
            message.Yes | message.No, message.No
            )
        
        if answer == message.Yes:
            self.save()
            self.MainWindow.close()

    def update(self):
        index = 0
        row = self.ui.tableWidget.rowCount()
        while sheet.max_row - 1> row:
            if sheet['A' + str(row + 1)].value:
                self.ui.tableWidget.insertRow(row)
                row = self.ui.tableWidget.rowCount()
            
            else:
                break

            
        items = [self.ui.task.itemText(i) for i in range(self.ui.task.count())]
        for row in range(1, sheet.max_row + 1):
            if sheet['B' + str(row)].value == 'not yet':
                task = sheet['A' + str(row)].value
                if task not in items: self.ui.task.addItem(task)
    
            for col in range(1, 5):
                cell = sheet.cell(row + 1, col).value
                if cell == None:
                    self.ui.tableWidget.removeRow(index)
                    break

                if col == 2:
                    status_box = QtWidgets.QComboBox()
                    if cell == 'Done':
                        status_box.addItems(['Done', 'not yet'])
                    else:
                        status_box.addItems(['not yet', 'Done'])
                    
                    status_box.setObjectName('status')
                    status_box.setProperty('row', index)
                    status_box.currentIndexChanged.connect(self.edit)
                    self.ui.tableWidget.setCellWidget(index, 1, status_box)
                
                else:
                    self.ui.tableWidget.setItem(index, col - 1, QtWidgets.QTableWidgetItem(str(cell)))

            index += 1

    def edit(self, item):

        try:
            row = item.row() + 2
            col = item.column() + 1
            data = item.text()

        except AttributeError:
            combo = self.sender()
            row = combo.property('row') + 2
            col = 2
            data = combo.currentText()
            sheet.cell(row, col).value = data


        if col != 2:
            sheet.cell(row, col).value = data
            
        items = [self.ui.task.itemText(i) for i in range(self.ui.task.count())]
        for row in range(1, sheet.max_row + 1):
            if sheet['B' + str(row)].value == 'not yet':
                task = sheet['A' + str(row)].value
                if task not in items: self.ui.task.addItem(task)
        
        wb.save('data.xlsx')
        

    
    def counterOption(self):
        button = self.sender()
        if button.text() == 'START':
            self.ui.status.setDisabled(True)
            self.timer.start(1000)
            self.ui.task.setDisabled(True)
            button.setText('STOP')
        
        else:
            self.timer.stop()
            self.ui.status.setEnabled(True)
            self.ui.task.setEnabled(True)
            self.timer.stop()
            button.setText('START')
            if self.loop%2: self.save()
                

        
    def taskOption(self):
        task = self.sender()
        self.task = task.currentText()
        

    def handleTimer(self):
        self.counter -= 1
        self.passed_time -= 1
        min = self.counter//60
        sec = int(self.counter - min * 60)
        self.ui.progressBar.setFormat(f'{min}:{sec}')
        value = self.ui.progressBar.m_value
        if value < self.curent_time:
            value = value + 1
            self.ui.progressBar.setValue(value)

        else:
            self.alert.play()
            self.loop += 1
            if self.loop%4 == 0:
                self.save()
                self.counter = 15 * 60
                self.curent_time = self.counter
                self.notify('Rest Time', 'You can take 15 minutes to rest.')
                
            elif self.loop%2 == 0:
                self.save()
                self.counter = 5 * 60
                self.curent_time = self.counter
                self.notify('Rest time', 'You can take 5 minutes to rest.')
                
            else:
                self.counter = 25 * 60
                self.curent_time = self.counter
                self.notify('Work time', f'Time to work on{self.task}')

            self.ui.progressBar.setValue(0)
            self.ui.progressBar.setRange(0, self.curent_time)
                
    
    def save(self, status='not yet'):

        expiry = QtCore.QDateTime.currentDateTime().toString(
            'yyyy/MM/dd hh:mm:ss'
            )
        time = str(timedelta(seconds=(25 * 60 - self.passed_time)))
        min = (25 * 60 - self.passed_time) // 60; sec = int((25 * 60 - self.passed_time) - min * 60)
        if sheet.max_row == 1:
            sheet['A2'] = self.task
            sheet['B2'] = status
            sheet['C2'] = expiry
            sheet['D2'] = time

        for row in range(2, sheet.max_row + 1):
            if sheet['A' + str(row)].value == self.task:
                sheet['B' + str(row)] = status
                sheet['C' + str(row)] = expiry

                duration = sheet['D' + str(row)].value

                try:
                    duration = re.compile(r'(\d*)\s(day|days),\s(\d*:\d*:\d*)').search(str(sheet['D' + str(row)].value))
                    days = int(duration.group(1))
                    duration = duration.group(3)

                except AttributeError:
                    days = 0
                    duration = str(sheet['D' + str(row)].value)

                lastTime = [int(i) for i in duration.split(':')]
                lastTime = (lastTime[0] + days * 24) * 3600 +\
                    (lastTime[1] + min) * 60 + lastTime[2] + sec

                sheet['D' + str(row)] = str(timedelta(seconds=lastTime))

                break
            elif row == sheet.max_row:
                sheet['A' + str(row)] = self.task
                sheet['B' + str(row)] = status
                sheet['C' + str(row)] = expiry
                sheet['D' + str(row)] = time

        self.passed_time = 25 * 60
        
        return self.update()

    def closeEvent(self, event):
        message = QtWidgets.QMessageBox.question(self, 'Quit', 'You sure you want to exit.',
        QtWidgets.QMessageBox.Yes |QtWidgets.QMessageBox.Cancel, QtWidgets.QMessageBox.Cancel)

        if message == QtWidgets.QMessageBox.Yes:
            self.save()
            event.accept()
        
        else:
            event.igonre()
        
    def notify(self, title, message):
        notification.notify(title=title, message=message,
        app_icon='logo.ico', timeout=5)

if __name__ == "__main__":
    import sys
    while True:
        try: # Try if there is data.xlsx file
            wb = load_workbook('data.xlsx')
            sheet = wb.active # Select sheet
            sheet['A1'] = 'Tasks'
            sheet['B1'] = 'Stauts'
            sheet['C1'] = 'Expiry Date'
            sheet['D1'] = 'Duration Taken'
            break # Stop the loop if data.xlsx is exist

        except: # Except if data.xlsx not exist
            wb = Workbook('data.xlsx') # Create data.xlsx file
            wb.create_sheet('data') # Create sheet
            wb.save('data.xlsx') # Save the data.xlsx

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    MainWindow.setStyleSheet(style_sheet)
    window = POMODORO(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())